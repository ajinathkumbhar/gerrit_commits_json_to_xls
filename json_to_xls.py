import json
from openpyxl import Workbook
import codecs

xls_file = "commits.xlsx"
commit_sheet = "commit_list"
title = "Commit list for module "

wb = Workbook()
ws = wb.create_sheet(title,0)

def finish():
    wb.save(xls_file)

def process_json():
    try:
        j_file = codecs.open("commit_list.json",'rU','utf-8')
    except IOError as err:
        print 'failed to open json file ' + str(err)
    for line in j_file:
        print line
        data = json.loads(line)
        print type(data)

        x_row = 2
        y_col = 2

        title_Cell = ws.cell(row=x_row ,column=y_col, value=title)
        x_row += 1
        for key, val in data.iteritems():
            print str(key) + " " + str(val)


        break
    # commit_data = json.load(j_file)
    # print commit_data
    # print data

def main():
    process_json()
    finish()




if __name__ == '__main__':
    main()